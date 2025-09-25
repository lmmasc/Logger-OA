"""
Funciones para extraer datos de operadores desde archivos Excel (.xlsx).
Específicamente diseñado para detectar y procesar formato chileno.
"""

import openpyxl
from datetime import datetime
import sys
import os
import re

# Agregar el directorio src al path para imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../../.."))

from utils.text import normalize_ascii
from utils.resources import get_resource_path
from domain.callsign_utils import callsign_to_country


def _extract_cutoff_date_from_filename(excel_path):
    """
    Extrae la fecha de cutoff del nombre del archivo Excel.
    Busca patrones de fecha en formato mes-día-año.

    Args:
        excel_path (str): Ruta al archivo Excel

    Returns:
        int: Timestamp de la fecha de cutoff, o fecha actual si no se puede extraer
    """
    filename = os.path.basename(excel_path)

    # Patrones para buscar fechas: mes-día-año en diferentes formatos
    patterns = [
        r"(\d{1,2})[_\-](\d{1,2})[_\-](\d{4})",  # MM-DD-YYYY o MM_DD_YYYY
        r"(\d{1,2})(\d{2})(\d{4})",  # MMDDYYYY
        r"(\d{2})(\d{2})(\d{4})",  # MMDDYYYY (con ceros)
    ]

    for pattern in patterns:
        match = re.search(pattern, filename)
        if match:
            try:
                month = int(match.group(1))
                day = int(match.group(2))
                year = int(match.group(3))

                # Validar rangos básicos
                if 1 <= month <= 12 and 1 <= day <= 31 and 2000 <= year <= 2100:
                    cutoff_date = datetime(year, month, day)
                    return int(cutoff_date.timestamp())
            except (ValueError, OverflowError):
                continue

    # Si no se puede extraer, usar fecha actual
    return int(datetime.now().timestamp())


def extract_operators_from_excel(excel_path):
    """
    Extrae los datos de operadores desde el archivo Excel especificado.
    Detecta automáticamente si es un formato chileno y procesa los datos correctamente.

    Args:
        excel_path (str): Ruta al archivo Excel (.xlsx)

    Returns:
        list: Lista de diccionarios con los datos de operadores

    Raises:
        Exception: Si el formato no es compatible o no es chileno
    """
    results = []
    seen_callsigns = set()
    excel_path = get_resource_path(excel_path)

    # Extraer fecha de cutoff del nombre del archivo
    cutoff_timestamp = _extract_cutoff_date_from_filename(excel_path)

    try:
        workbook = openpyxl.load_workbook(excel_path, read_only=True)
        worksheet = workbook.active

        if not worksheet or worksheet.max_row < 2:
            raise Exception("El archivo Excel está vacío o no tiene datos")

        # Leer la primera fila para detectar formato
        header_row = next(worksheet.iter_rows(values_only=True, max_row=1))
        if not header_row:
            raise Exception("No se pudo leer la cabecera del Excel")

        # Normalizar cabeceras para comparación
        normalized_headers = [str(h).strip().lower() if h else "" for h in header_row]

        # Verificar si es formato chileno
        if not _is_chilean_format(normalized_headers):
            raise Exception(
                "El formato del Excel no corresponde a operadores chilenos. "
                "Se esperan columnas: Licencia, Señal Distintiva, Nombre, Rut, Región, Comuna, Fecha Vencimiento"
            )

        print("✓ Formato chileno detectado correctamente")

        # Mapear columnas del formato chileno
        column_mapping = _map_chilean_columns(normalized_headers)

        # Procesar filas de datos
        row_count = 0
        for row_data in worksheet.iter_rows(values_only=True, min_row=2):
            row_count += 1
            if not any(cell is not None and str(cell).strip() for cell in row_data):
                continue  # Saltar filas vacías

            try:
                operator_data = _process_chilean_row(
                    row_data, column_mapping, cutoff_timestamp
                )
                if operator_data and operator_data["callsign"] not in seen_callsigns:
                    results.append(operator_data)
                    seen_callsigns.add(operator_data["callsign"])
            except Exception as e:
                print(f"⚠ Error procesando fila {row_count + 1}: {e}")
                continue

        workbook.close()
        print(f"✓ Procesados {len(results)} operadores chilenos únicos")

    except Exception as e:
        if "workbook" in locals():
            workbook.close()
        raise Exception(f"Error al procesar archivo Excel: {str(e)}")

    return results


def _is_chilean_format(headers):
    """
    Verifica si las cabeceras corresponden al formato chileno.

    Args:
        headers (list): Lista de cabeceras normalizadas

    Returns:
        bool: True si es formato chileno
    """
    required_chilean_headers = [
        "licencia",
        "señal distintiva",
        "nombre",
        "rut",
        "región",
        "comuna",
        "fecha vencimiento",
    ]

    # Verificar que todas las cabeceras requeridas estén presentes
    headers_joined = " ".join(headers).lower()
    for required in required_chilean_headers:
        if required not in headers_joined:
            return False

    return True


def _map_chilean_columns(headers):
    """
    Mapea las columnas del formato chileno a índices.

    Args:
        headers (list): Lista de cabeceras normalizadas

    Returns:
        dict: Mapeo de campo -> índice de columna
    """
    mapping = {}

    for i, header in enumerate(headers):
        header_clean = header.strip().lower()
        if "licencia" in header_clean:
            mapping["licencia"] = i
        elif "señal" in header_clean and "distintiva" in header_clean:
            mapping["callsign"] = i
        elif "nombre" in header_clean and "rut" not in header_clean:
            mapping["nombre"] = i
        elif "rut" in header_clean:
            mapping["rut"] = i
        elif "región" in header_clean or "region" in header_clean:
            mapping["region"] = i
        elif "comuna" in header_clean:
            mapping["comuna"] = i
        elif "fecha" in header_clean and "vencimiento" in header_clean:
            mapping["fecha_vencimiento"] = i

    return mapping


def _process_chilean_row(row_data, column_mapping, cutoff_timestamp):
    """
    Procesa una fila de datos chilenos y la convierte al formato interno.

    Args:
        row_data (tuple): Datos de la fila
        column_mapping (dict): Mapeo de columnas
        cutoff_timestamp (int): Timestamp de la fecha de cutoff del archivo

    Returns:
        dict: Datos del operador en formato interno
    """
    # Extraer datos usando el mapeo
    licencia = _get_cell_value(row_data, column_mapping.get("licencia"))
    callsign = _get_cell_value(row_data, column_mapping.get("callsign"))
    nombre = _get_cell_value(row_data, column_mapping.get("nombre"))
    rut = _get_cell_value(row_data, column_mapping.get("rut"))
    region = _get_cell_value(row_data, column_mapping.get("region"))
    comuna = _get_cell_value(row_data, column_mapping.get("comuna"))
    fecha_venc = (
        row_data[column_mapping["fecha_vencimiento"]]
        if column_mapping.get("fecha_vencimiento") is not None
        else None
    )

    # Validar callsign obligatorio
    if not callsign:
        raise Exception("Callsign vacío")

    # Normalizar callsign: QUITAR ESPACIOS y convertir a mayúsculas
    callsign_clean = callsign.strip().replace(" ", "").upper()

    # Verificar que el callsign es chileno
    country_code = callsign_to_country(callsign_clean)
    if country_code != "CHL":
        raise Exception(
            f"Callsign {callsign_clean} no es chileno (detectado: {country_code})"
        )

    # Procesar fecha de vencimiento
    expiration_timestamp = None
    if fecha_venc:
        expiration_timestamp = _convert_date_to_timestamp(fecha_venc)

    # Construir región normalizada: "REGION - COMUNA"
    region_normalized = _build_normalized_region(region, comuna)

    # Construir datos del operador en formato interno
    operator_data = {
        "callsign": callsign_clean,
        "name": nombre.strip() if nombre else "",
        "category": "",  # No disponible en formato chileno
        "type": "",  # No disponible en formato chileno
        "region": region_normalized,
        "district": "",  # No aplica para Chile
        "province": "",  # No aplica para Chile
        "department": "",  # No aplica para Chile
        "license": licencia.strip() if licencia else "",
        "resolution": rut.strip() if rut else "",  # RUT en lugar de resolución
        "expiration_date": expiration_timestamp,
        "cutoff_date": cutoff_timestamp,  # Fecha del archivo
        "country": "CHL",  # Código ITU de Chile
        "enabled": 1,  # Por defecto habilitado
    }

    return operator_data


def _get_cell_value(row_data, column_index):
    """
    Obtiene el valor de una celda de forma segura.

    Args:
        row_data (tuple): Datos de la fila
        column_index (int|None): Índice de columna

    Returns:
        str: Valor de la celda como string
    """
    if column_index is None or column_index >= len(row_data):
        return ""

    cell_value = row_data[column_index]
    if cell_value is None:
        return ""

    return str(cell_value).strip()


def _convert_date_to_timestamp(date_value):
    """
    Convierte un valor de fecha a timestamp UTC.

    Args:
        date_value: Valor de fecha (datetime, string, etc.)

    Returns:
        int|None: Timestamp UTC o None si no se puede convertir
    """
    if isinstance(date_value, datetime):
        return int(date_value.timestamp())

    if isinstance(date_value, str) and date_value.strip():
        # Intentar parsear diferentes formatos de fecha
        formats = ["%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"]

        for fmt in formats:
            try:
                dt = datetime.strptime(date_value.strip(), fmt)
                return int(dt.timestamp())
            except ValueError:
                continue

    return None


def _build_normalized_region(region, comuna):
    """
    Construye la región normalizada en formato "REGION - COMUNA".

    Args:
        region (str): Región chilena
        comuna (str): Comuna chilena

    Returns:
        str: Región normalizada
    """
    if not region and not comuna:
        return ""

    region_clean = normalize_ascii(region or "").upper()
    comuna_clean = normalize_ascii(comuna or "").upper()

    if region_clean and comuna_clean:
        return f"{region_clean} - {comuna_clean}"
    elif region_clean:
        return region_clean
    elif comuna_clean:
        return comuna_clean
    else:
        return ""
