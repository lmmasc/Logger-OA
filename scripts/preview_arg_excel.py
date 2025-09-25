import openpyxl

wb = openpyxl.load_workbook(
    "BaseDocs/ARG - Listado de Radioaficionado 12.12.2024.xlsx", read_only=True
)
ws = wb.active

print("Cabeceras:")
headers = [
    str(cell).strip() for cell in next(ws.iter_rows(values_only=True, max_row=1))
]
print(headers)

print("\nPrimeras filas:")
for i, row in enumerate(ws.iter_rows(values_only=True, min_row=2), 1):
    print(row)
    if i >= 5:
        break
wb.close()
